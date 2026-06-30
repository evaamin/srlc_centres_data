"""
Event Data Cleaning & Consolidation Script
--------------------------------------------
Purpose:
    The raw export has 1 event spread across MULTIPLE rows. The "extra" rows
    look blank but actually carry additional values for two repeating field
    groups: Beneficiary Details and Items Donated (an event can have more
    than one beneficiary / item, and the export gives each one its own row).

    This script:
      1. Detects the real event rows (non-blank "Centre" column) and groups
         every following blank row under that event.
      2. Merges the repeating Beneficiary Details / Items Donated values into
         a single cell per event (joined with "; ") so no data is lost.
      3. Drops the duplicate, always-empty "Brief Description of Event" column.
      4. Strips HTML tags/entities from ALL description-style columns:
         "Brief Description of Event.1", "Awards / Accolades Awarded during
         event", "Written Testimonials", and "Additonal Details".
      5. Adds a few starter helper/QC columns (clearly marked - update after
         discussing with Nidhi).
      6. SAFETY CHECK: scans every future file for any column that
         unexpectedly has data in a blank row but ISN'T in REPEATING_COLUMNS.
         If found, prints a warning so you know to review/update the script
         instead of silently losing data.
      7. Converts "No of lives touched" to a real number (float) instead of
         text, so it can be summed/charted directly in Excel.
      8. Adds a new "Centre - Geo" column: the Centre value with org labels
         (SRLC, Youth, Central, state codes) stripped out, leaving just the
         city/place name. If NO city can be extracted from "Centre" at all
         (e.g. "Environmental Care"), falls back to that row's "Centre/City"
         column value instead; if that's also empty, keeps the original
         Centre text unchanged.

Usage:
    python clean_events.py input.csv output.xlsx
"""

import sys
import re
import pandas as pd
from bs4 import BeautifulSoup

# Columns that repeat across multiple rows for the same event.
# These get CONCATENATED (joined by "; ") instead of just taking the first value.
REPEATING_COLUMNS = [
    "Beneficiary Details/Beneficiary",
    "Beneficiary Details/Beneficiary Location",
    "Beneficiary Details/Beneficiary Name",
    "Beneficiary Details/Beneficiary Type",
    "Items Donated/Display Name",
    "Items Donated/Items",
    "Items Donated/Items Donated",
    "Items Donated/Quantity/Weight",
    "Items Donated/Unit",
]

# The column that actually holds event-identifying data on the "primary" row.
ANCHOR_COLUMN = "Centre"

# The two duplicate description columns in the raw export.
EMPTY_DESC_COLUMN = "Brief Description of Event"      # always empty in this export
HTML_DESC_COLUMN = "Brief Description of Event.1"      # the real, HTML-formatted text

# All columns (besides HTML_DESC_COLUMN above) that also contain HTML formatting
# and need the same tag-stripping treatment.
OTHER_HTML_COLUMNS = [
    "Awards / Accolades Awarded during event",
    "Written Testimonials",
    "Additonal Details",
]

# Columns that should be converted to real numbers (not text) in the output,
# so Excel can sum/average/chart them directly.
# NOTE: "Centre/ID" and "Parent Zone/ID" are kept as text (identifiers, not quantities).
NUMERIC_COLUMNS = [
    "No of lives touched",
    "Cost",
    "Total Amount Donated by First Time Donors",
    "Number of Sevaks",
    "Number of First-Time SRLC Sevaks",
    "Number of Seva Hours Total",
    "Items Donated/Quantity/Weight",
    "Total Number of Kits/Servings (if applicable)",
    "Number of Attendees",
    "Number of Attendees (WAT)",
    "Number of Attendees (AC)",
    "Duration of Class (Hours)",
    "Number of Patients",
    "Number of Procedures",
    "Number of individuals screened",
    "Units of blood collected",
    "mL of blood per unit",
    "Number of New Registrants",
    "Number of Individuals Supported",
    "Number of Trees Planted",
    "Weight of Garbage Collected",
]

# Columns that represent duration in hours and should be formatted as [h]:mm
# in Excel (e.g. 250 hours -> 250:00), right-aligned, and SUM-compatible.
# Values are converted from plain hours (e.g. 20) to Excel time fractions (20/24).
HOURS_COLUMNS = [
    "Number of Seva Hours Total",
    "Duration of Class (Hours)",
]

# Words to strip out of "Centre" when building the new "Centre - Geo" column.
# These are organizational/branch labels, not part of the city name.
CENTRE_NOISE_WORDS = ["SRLC", "Youth", "Central"]


def extract_centre_geo(centre: str, fallback_city: str = "") -> str:
    """
    Build a clean "city only" value from the "Centre" column.

    Examples:
        "Atlanta - GA SRLC"          -> "Atlanta"
        "Manchester - Central SRLC"  -> "Manchester"
        "Ghatkopar - Youth SRLC"     -> "Ghatkopar"
        "Environmental Care"        -> falls back to the "Centre/City" column
                                         value for that row (per manager's
                                         instruction: "if Centre does not
                                         contain a city, keep what is there
                                         in the city column")
        "SRA USA"                   -> no fallback available -> kept as-is

    Logic:
      1. Try to strip known noise words/patterns: "SRLC", "Youth", "Central",
         and any standalone 2-letter US state code (e.g. "GA", "TX", "NY").
      2. If nothing was found/stripped at all -> this Centre name has no
         recognizable city pattern -> use the fallback_city value (the row's
         "Centre/City" column) if it has something in it; otherwise keep the
         original Centre text unchanged.
    """
    original = str(centre).strip()
    if original == "":
        return ""

    fallback = str(fallback_city).strip()

    text = original
    changed = False

    # Remove noise words (SRLC, Youth, Central) wherever they appear as whole words
    for word in CENTRE_NOISE_WORDS:
        pattern = re.compile(r"\b" + re.escape(word) + r"\b", re.IGNORECASE)
        if pattern.search(text):
            text = pattern.sub("", text)
            changed = True

    # Remove standalone 2-letter uppercase state codes (e.g. "GA", "TX", "NY")
    state_code_pattern = re.compile(r"(?<![A-Za-z])[A-Z]{2}(?![A-Za-z])")
    if state_code_pattern.search(text):
        text = state_code_pattern.sub("", text)
        changed = True

    if not changed:
        # Nothing recognizable to strip (e.g. "Environmental Care", "SRA USA")
        # -> fall back to the Centre/City column if it has a value, otherwise
        # keep the original Centre text unchanged.
        return fallback if fallback != "" else original

    # Tidy up leftover dashes/extra whitespace left behind after removal
    text = re.sub(r"-{2,}", "-", text)
    text = re.sub(r"\s*-\s*", " - ", text)
    text = text.strip(" -")
    text = re.sub(r"\s+", " ", text).strip()

    if text == "":
        # Stripped everything away and nothing useful is left (e.g. a Centre
        # that was JUST "TX SRLC" with no city at all) -> fall back too.
        return fallback if fallback != "" else original

    return text


def clean_html(text: str) -> str:
    """Strip HTML tags/entities and tidy whitespace from a description field."""
    if not isinstance(text, str) or text.strip() == "":
        return ""
    soup = BeautifulSoup(text, "html.parser")
    # Use newlines so paragraph breaks aren't lost when tags are stripped
    raw = soup.get_text(separator="\n")
    # Collapse repeated blank lines / spaces left behind by empty <p><br></p> tags
    lines = [line.strip() for line in raw.splitlines()]
    lines = [line for line in lines if line]
    cleaned = "\n".join(lines)
    # Remove stray tab/zero-width characters sometimes embedded in the HTML
    cleaned = re.sub(r"[\t\u200b]", "", cleaned)
    return cleaned.strip()


def check_for_unexpected_data(df: pd.DataFrame) -> list:
    """
    SAFETY CHECK for future files.

    This script assumes blank-"Centre" rows ONLY ever carry extra data in
    the columns listed in REPEATING_COLUMNS. If a future export adds a new
    column that also splits across rows (something we haven't seen before),
    that data would otherwise be silently dropped.

    This function scans every blank-Centre row and flags any column NOT in
    REPEATING_COLUMNS that unexpectedly has a value. Returns a list of
    warning messages (empty list = nothing unexpected found).
    """
    warnings = []
    is_blank_row = df[ANCHOR_COLUMN].astype(str).str.strip() == ""
    blank_rows = df[is_blank_row]

    # Columns we don't expect any value in in a blank row.
    # (Everything except the known repeating columns and the anchor column itself.)
    columns_to_check = [c for c in df.columns if c not in REPEATING_COLUMNS and c != ANCHOR_COLUMN]

    for col in columns_to_check:
        nonblank_count = (blank_rows[col].astype(str).str.strip() != "").sum()
        if nonblank_count > 0:
            warnings.append(
                f"⚠️  Found {nonblank_count} blank-row value(s) in column "
                f"'{col}' — this column is NOT in REPEATING_COLUMNS, so its "
                f"data may be getting silently ignored. Review this column "
                f"and add it to REPEATING_COLUMNS at the top of the script "
                f"if it should be merged in."
            )
    return warnings


def consolidate(df: pd.DataFrame) -> pd.DataFrame:
    df = df.fillna("")

    # Mark every row that starts a new event (non-blank anchor column)
    is_primary = df[ANCHOR_COLUMN].astype(str).str.strip() != ""

    # Group ID: increments every time a new primary row is seen
    group_id = is_primary.cumsum()

    consolidated_rows = []
    for _, group in df.groupby(group_id):
        primary_row = group.iloc[0].copy()

        # --- Merge repeating columns across all rows in this event's group ---
        beneficiary_entry_count = 0
        item_entry_count = 0

        for col in REPEATING_COLUMNS:
            values = [str(v).strip() for v in group[col] if str(v).strip() != ""]
            primary_row[col] = "; ".join(values)

        # Count how many distinct beneficiary / item rows contributed data
        # (used for the helper QC columns below)
        beneficiary_entry_count = (group["Beneficiary Details/Beneficiary"].astype(str).str.strip() != "").sum()
        item_entry_count = (group["Items Donated/Items"].astype(str).str.strip() != "").sum()

        # --- Helper / QC columns (placeholders - refine after talking to Nidhi) ---
        primary_row["QC: Rows Merged Into This Event"] = len(group)
        primary_row["QC: Number of Beneficiary Entries"] = beneficiary_entry_count
        primary_row["QC: Number of Item Entries"] = item_entry_count
        primary_row["QC: Missing Description Flag"] = (
            clean_html(group.iloc[0][HTML_DESC_COLUMN]) == ""
        )

        consolidated_rows.append(primary_row)

    result = pd.DataFrame(consolidated_rows).reset_index(drop=True)
    return result


def main(input_path: str, output_path: str):
    df = pd.read_csv(input_path, dtype=str, keep_default_na=False)

    # --- SAFETY CHECK: warn if this file has a pattern we haven't seen before ---
    warnings = check_for_unexpected_data(df)
    if warnings:
        print("=" * 70)
        print("SAFETY CHECK WARNINGS - please review before trusting the output:")
        print("=" * 70)
        for w in warnings:
            print(w)
        print("=" * 70)
    else:
        print("Safety check passed: no unexpected data found in blank rows.")

    result = consolidate(df)

    # --- FIX 1: Clean HTML in the main description column (existing) ---
    result[HTML_DESC_COLUMN] = result[HTML_DESC_COLUMN].apply(clean_html)
    result = result.rename(columns={HTML_DESC_COLUMN: "Brief Description of Event (cleaned)"})

    # --- FIX 1 (continued): Clean HTML in the OTHER columns flagged by manager ---
    for col in OTHER_HTML_COLUMNS:
        if col in result.columns:
            result[col] = result[col].apply(clean_html)

    # Drop the duplicate always-empty description column
    if EMPTY_DESC_COLUMN in result.columns:
        result = result.drop(columns=[EMPTY_DESC_COLUMN])

    # --- FIX 2: Convert "Lives Touched" (and any other listed columns) to real numbers ---
    for col in NUMERIC_COLUMNS:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors="coerce")

    # --- FIX 3: Add the new "Centre - Geo" column right after "Centre" ---
    # Falls back to the "Centre/City" column when no city can be extracted
    # from "Centre" itself (e.g. "Environmental Care" -> uses its Centre/City
    # value), per manager's instruction.
    fallback_col = "Centre/City" if "Centre/City" in result.columns else None
    if fallback_col:
        result["Centre - Geo"] = [
            extract_centre_geo(c, fb) for c, fb in zip(result["Centre"], result[fallback_col])
        ]
    else:
        result["Centre - Geo"] = result["Centre"].apply(extract_centre_geo)
    cols = list(result.columns)
    cols.remove("Centre - Geo")
    centre_pos = cols.index("Centre") + 1
    cols.insert(centre_pos, "Centre - Geo")
    result = result[cols]

    if output_path.endswith(".xlsx"):
        result.to_excel(output_path, index=False)
        # Apply number formatting so Excel treats numeric columns as numbers,
        # not text — enabling SUM, AVERAGE, charting, etc. directly.
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        wb = load_workbook(output_path)
        ws = wb.active
        # Build a map from column header -> column index
        header_to_col = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

        # --- Standard numeric columns: plain number format ---
        for col_name in NUMERIC_COLUMNS:
            if col_name not in header_to_col:
                continue
            col_idx = header_to_col[col_name]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if cell.value is not None and str(cell.value).strip() != "":
                    try:
                        val = float(cell.value)
                        cell.value = val
                        # Integer format for whole numbers, 2 decimals otherwise
                        cell.number_format = "#,##0" if val == int(val) else "#,##0.00"
                    except (ValueError, TypeError):
                        pass  # leave non-numeric cells as-is

        # --- Hours columns: [h]:mm format, right-aligned ---
        # Excel stores time as a fraction of a day, so we divide hours by 24.
        # The [h]:mm format shows total elapsed hours (e.g. 250:00, 20:30).
        # SUM still works correctly with this format.
        for col_name in HOURS_COLUMNS:
            if col_name not in header_to_col:
                continue
            col_idx = header_to_col[col_name]
            # Right-align the header too
            ws.cell(1, col_idx).alignment = Alignment(horizontal="right")
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if cell.value is not None and str(cell.value).strip() != "":
                    try:
                        hours = float(cell.value)
                        cell.value = hours / 24   # convert to Excel time fraction
                        cell.number_format = "[h]:mm"
                        cell.alignment = Alignment(horizontal="right")
                    except (ValueError, TypeError):
                        pass  # leave non-numeric cells as-is

        wb.save(output_path)
    else:
        result.to_csv(output_path, index=False)

    print(f"Input rows:  {len(df)}")
    print(f"Output rows: {len(result)}  (1 row per event)")
    print(f"Saved to: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python clean_events.py <input.csv> <output.xlsx>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
